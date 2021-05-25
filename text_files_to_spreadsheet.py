#! python3
# text_files_to_spreadsheet.py - reads several text files and put them on a xlsx file
# Use: text_files_to_spreadsheet.py myfile1.txt myfile2.txt [...]
import openpyxl, sys

wb = openpyxl.Workbook()
ws = wb.active

filenames = sys.argv[1:]
for i in range(len(filenames)):
    filename = filenames[i]
    column   = openpyxl.utils.get_column_letter(i + 1)
    row = 1
    
    f = open(filename, 'r')
    for line in f.readlines():
        ws[column + str(row)] = line
        row += 1

wb.save('spreasheaded.xlsx') # really?
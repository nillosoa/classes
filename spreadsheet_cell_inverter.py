#! python3
# spreadsheet_cell_inverter.py - inverts all rows and cells of an excel file

import openpyxl, sys

filename = sys.argv[1]
wb = openpyxl.load_workbook(filename)
ws = wb.active
title = ws.title

x = [] # cordinates
y = [] # values
# all values and coordinates will have the same index
for colObj in ws.columns:
    for celObj in colObj:
        x.append(celObj.coordinate)
        y.append(celObj.value)

new_wb = openpyxl.Workbook()
new_ws = new_wb.active
new_ws.title = title

for i in range(len(x)):
    coordinate  =  x[i]
    value       =  y[i]
    col, row = openpyxl.utils.coordinate_from_string(coordinate) # Transforms 'A1' into the tuple ('A', 1)
    col_idx  = openpyxl.utils.column_index_from_string(col) # Transforms A(col, column) into a number(row), so do speak
    row_let  = openpyxl.utils.get_column_letter(row) # Transforms 1(row) into a column(A)
    new_coordinate = row_let + str(col_idx) # Concatenates the values
    new_ws[new_coordinate] = value # Put it into our new worksheet
    #print(new_coordinate, coordinate, value) # Debugging HeHe

new_wb.save(filename.split('.')[0] + '_inverted.xlsx')

# NOTE, it seens openpyxl has a limit of 18279 columns per xlsx file,
# so the script may crash with files with more than 18279 rows